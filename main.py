import time
import glob
import io
import urllib3

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

from libs.lib_sessions import create_request_session
from parser_yandex_pairs import YandexPairParser


def insert_image(ws, image: Image, column: str):
    ws.row_dimensions[ws.max_row].height = 100
    ratio = image.width / image.height
    image.height = 100
    image.width = int(ratio * 100.0)
    ws.add_image(image, f'{column}{ws.max_row}')


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 20
    ws.append(['Исходное изображение', 'Найденное изображение', 'Название', 'URL'])

    parser = YandexPairParser(settings={}, request_session=create_request_session({}))
    http = urllib3.PoolManager()

    t1 = time.time()
    for img_path in glob.glob(r'Сжатые\Сжатые\*'):
        src_image = openpyxl.drawing.image.Image(img_path)
        found_pages = parser.execute(img_path)
        for page in found_pages:
            print(page)
            ws.append(['', '', page.page.title, page.page.url])
            insert_image(ws, src_image, 'A')

            try:
                r = http.request('GET', page.image.url)
                match_img = Image(io.BytesIO(r.data))
                if match_img.format == 'webp':
                    continue

                insert_image(ws, match_img, 'B')
            except Exception as e:
                print('Exception while downloading img: ', e)

        # print(output)
        # print('Size:', len(output))
        break

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = cell.alignment.copy(wrapText=True)

    wb.save('test.xlsx')

    t2 = time.time()
    print('Time:', t2 - t1)


if __name__ == '__main__':
    main()
